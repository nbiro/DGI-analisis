import requests
import pandas as pd
import yfinance as yf
import openpyxl

def download_and_save(url, filename):
    r = requests.get(url, allow_redirects=True)
    open(filename, 'wb').write(r.content)


us_dividend_champions_url = 'https://bit.ly/USDividendChampions1'
us_dividend_champions_file_name = 'resources/us_dividend_champions.xlsx'

download_and_save(us_dividend_champions_url, us_dividend_champions_file_name)



def get_ticker_price(tickers):
  data = yf.download(tickers, period ='1d')
  return data

wb = openpyxl.load_workbook(us_dividend_champions_file_name)

tickers = []
for row in wb['All'].iter_rows(min_row=4):
  tickers.append(row[0].value)
  
print('Downloading ticker data')
tickers_data = get_ticker_price(tickers)

for row in wb['All'].iter_rows(min_row=4):
  ticker = row[0].value
  try:
    price = float(tickers_data['High'][ticker][-1:])
    row[5].value = price
    print(f"Updating {ticker} with price {price}")
  except KeyError:
    row[5].value = "NaN"
    print(f"No data available for {ticker}, or regular price is unknown")
    price = None
    
  current_div = row[8].value
  payout_per_year = row[9].value
  
  if current_div and payout_per_year and price:
    div_yield = (current_div * payout_per_year) / price * 100
    row[6].value = div_yield

wb.save(us_dividend_champions_file_name)